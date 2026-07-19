"""
Excel-Export mit openpyxl.

Zwei Blaetter:
  * "Aushang"     - farbiges Wochenraster (Zeilen = Mitarbeiter, Spalten = Tage)
  * "Abrechnung"  - Stunden je Mitarbeiter/Tag mit Summe, Soll und Differenz
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import db

WD = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
MONTHS = ["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
KIND_MARK = {"kurs": "Kurs", "telefon": "Tel", "zbv": "zbV", "event": "Termin"}

THIN = Side(style="thin", color="D6DBE2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY = "1B2A41"


def _text_color(hex_color: str) -> str:
    h = (hex_color or "#888888").lstrip("#")
    r, g, b = (int(h[i:i + 2], 16) / 255 for i in (0, 2, 4))
    lin = lambda c: c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4
    L = 0.2126 * lin(r) + 0.7152 * lin(g) + 0.0722 * lin(b)
    return "FF12212F" if L > 0.5 else "FFF2F6FB"


def _fill(hex_color: str) -> PatternFill:
    c = (hex_color or "#888888").lstrip("#").upper()
    return PatternFill("solid", fgColor="FF" + c)


def _min(t: str) -> int:
    h, m = t.split(":")
    return int(h) * 60 + int(m)


def _gather(location_id, year, week):
    conn = db.connect()
    try:
        loc = dict(conn.execute("SELECT * FROM locations WHERE id=?", (location_id,)).fetchone())
        emps = [dict(e) for e in conn.execute(
            "SELECT * FROM employees WHERE location_id=? AND active=1 ORDER BY sort, id", (location_id,))]
        rooms = {r["id"]: r["name"] for r in conn.execute("SELECT * FROM rooms WHERE location_id=?", (location_id,))}
        assigns = [dict(r) for r in conn.execute(
            "SELECT * FROM assignments WHERE location_id=? AND iso_year=? AND iso_week=? ORDER BY start_time",
            (location_id, year, week))]
        emp_ids = [e["id"] for e in emps]
        absences = []
        if emp_ids:
            ph = ",".join("?" * len(emp_ids))
            absences = [dict(r) for r in conn.execute(
                f"SELECT * FROM absences WHERE iso_year=? AND iso_week=? AND employee_id IN ({ph})",
                (year, week, *emp_ids))]
        settings = {r["key"]: db.loads(r["value"]) for r in conn.execute("SELECT * FROM settings")}
        dates = [date.fromisocalendar(year, week, d).isoformat() for d in range(1, 8)]
        return loc, emps, rooms, assigns, absences, settings, dates
    finally:
        conn.close()


def _abs_on(absences, emp_id, iso, abs_cat):
    out = []
    for ab in absences:
        if ab["employee_id"] != emp_id:
            continue
        hit = (not ab["date"]) or (ab.get("end_date") and ab["date"] <= iso <= ab["end_date"]) or (ab["date"] == iso)
        if hit:
            lbl = abs_cat.get(ab["type"], {}).get("label", ab["type"])
            per = ab.get("period")
            out.append(lbl + (f" ({per})" if per and per != "ganztags" else ""))
    return out


def export(location_id: int, year: int, week: int):
    loc, emps, rooms, assigns, absences, settings, dates = _gather(location_id, year, week)
    abs_cat = {t["id"]: t for t in (settings.get("absence_types") or [])}
    d0 = date.fromisoformat(dates[0]); d6 = date.fromisoformat(dates[6])
    rng = f"{d0.day}.–{d6.day}. {MONTHS[d6.month]} {d6.year}"

    wb = Workbook()

    # ---------------- Blatt 1: Aushang -----------------------------------
    ws = wb.active
    ws.title = "Aushang"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Dienstplan · {loc['name']} · KW {week} · {rng}"
    ws["A1"].font = Font(bold=True, size=14, color="FF" + NAVY)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 24

    hdr_fill = PatternFill("solid", fgColor="FF" + NAVY)
    hdr_font = Font(bold=True, color="FFF2F6FB")
    ws.cell(3, 1, "Mitarbeiter").fill = hdr_fill
    ws.cell(3, 1).font = hdr_font
    for i, iso in enumerate(dates):
        d = date.fromisoformat(iso)
        c = ws.cell(3, 2 + i, f"{WD[i][:2]} {d.day}.{d.month}.")
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    for col in range(1, 9):
        ws.cell(3, col).border = BORDER

    for r, emp in enumerate(emps, start=4):
        name_cell = ws.cell(r, 1, f"{emp.get('kuerzel') or ''}  {emp['name']}")
        name_cell.fill = _fill(emp.get("color"))
        name_cell.font = Font(bold=True, color=_text_color(emp.get("color")))
        name_cell.border = BORDER
        name_cell.alignment = Alignment(vertical="center")
        for i, iso in enumerate(dates):
            parts = []
            for a in assigns:
                if a["employee_id"] == emp["id"] and a["date"] == iso:
                    mark = KIND_MARK.get(a["kind"], "")
                    home = " ⌂" if a.get("work_mode") == "home" else ""
                    txt = f"{a['start_time']}–{a['end_time']}"
                    if mark:
                        txt += f" {mark}"
                    parts.append(txt + home)
            parts += _abs_on(absences, emp["id"], iso, abs_cat)
            cell = ws.cell(r, 2 + i, "\n".join(parts))
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            cell.border = BORDER
        ws.row_dimensions[r].height = 42

    ws.column_dimensions["A"].width = 26
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.freeze_panes = "B4"

    # ---------------- Blatt 2: Abrechnung --------------------------------
    ws2 = wb.create_sheet("Abrechnung")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = f"Stundenabrechnung · {loc['name']} · KW {week}"
    ws2["A1"].font = Font(bold=True, size=13, color="FF" + NAVY)
    ws2.merge_cells("A1:K1")

    headers = ["Mitarbeiter"] + [WD[i][:2] for i in range(7)] + ["Summe", "Soll", "Diff"]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(3, col, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")
        c.border = BORDER

    for r, emp in enumerate(emps, start=4):
        ws2.cell(r, 1, f"{emp.get('kuerzel') or ''} {emp['name']}").border = BORDER
        day_min = [0] * 7
        for a in assigns:
            if a["employee_id"] != emp["id"] or a["kind"] == "event":
                continue
            if a["date"] in dates:
                day_min[dates.index(a["date"])] += _min(a["end_time"]) - _min(a["start_time"])
        total = 0
        for i in range(7):
            h = day_min[i] / 60
            total += h
            cell = ws2.cell(r, 2 + i, round(h, 2) if h else None)
            cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
        soll = emp.get("weekly_hours") or 0
        ws2.cell(r, 9, round(total, 2)).font = Font(bold=True)
        ws2.cell(r, 9).border = BORDER; ws2.cell(r, 9).alignment = Alignment(horizontal="center")
        ws2.cell(r, 10, soll).border = BORDER; ws2.cell(r, 10).alignment = Alignment(horizontal="center")
        diff = round(total - soll, 2)
        dc = ws2.cell(r, 11, diff)
        dc.border = BORDER; dc.alignment = Alignment(horizontal="center")
        dc.font = Font(color="FFB00020" if diff < 0 else "FF1E7B34")

    ws2.column_dimensions["A"].width = 26
    for col in range(2, 12):
        ws2.column_dimensions[get_column_letter(col)].width = 8
    ws2.freeze_panes = "B4"

    buf = BytesIO()
    wb.save(buf)
    fname = f"Dienstplan_{loc.get('kuerzel') or loc['name']}_KW{week}.xlsx"
    return fname, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", buf.getvalue()
